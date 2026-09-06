import sys
import os
import csv
import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from engine.io_manager import IOManager

sys.stdout.reconfigure(encoding='utf-8')

io_mgr = IOManager()
dataset = io_mgr.load_dataset()
records = dataset.get("records", [])

print(f"Generating 10k MC Final export for {len(records)} records...")

all_rows = []
drug_rows = []
polymer_rows = []

for r in records:
    eid = r["entity_id"]
    etype = r.get("entity_type", "drug").lower()
    raw_name = r.get("name", "")
    abbrev = r.get("abbreviation") or ""
    
    sheet = io_mgr.generate_pharmapolyscope_ready_sheet(eid)
    fields_map = {f["key"]: f for f in sheet["fields"]}
    
    # Extract 10k MC Final values
    smiles_val = fields_map.get("canonical_smiles", {}).get("final_value") or fields_map.get("repeat_unit_smiles", {}).get("final_value")
    if isinstance(smiles_val, dict):
        smiles_val = smiles_val.get("value")
        
    mw_val = fields_map.get("mw", {}).get("final_value")
    if mw_val is None and etype == "polymer":
        ru = r.get("repeat_unit_smiles", {})
        mw_val = ru.get("repeat_unit_mw") if isinstance(ru, dict) else None
        
    tm_val = fields_map.get("tm_K", {}).get("final_value")
    tg_val = fields_map.get("tg_K", {}).get("final_value")
    dens_val = fields_map.get("density_g_cm3", {}).get("final_value")
    true_dens = r.get("density_g_cm3", {}).get("true_density_g_cm3") if isinstance(r.get("density_g_cm3"), dict) else None
    
    dD_val = fields_map.get("delta_D", {}).get("final_value")
    dP_val = fields_map.get("delta_P", {}).get("final_value")
    dH_val = fields_map.get("delta_H", {}).get("final_value")
    dt_val = fields_map.get("delta_t", {}).get("final_value")
    r0_val = fields_map.get("R0", {}).get("final_value", 7.5)
    
    logp_val = fields_map.get("logP", {}).get("final_value")
    tpsa_val = fields_map.get("TPSA", {}).get("final_value")
    hbd_val = fields_map.get("HBD", {}).get("final_value")
    hba_val = fields_map.get("HBA", {}).get("final_value")
    bcs_val = fields_map.get("BCS_class", {}).get("final_value")
    
    qc_status = sheet.get("qc_status", "APPROVED")
    
    # Unified row
    unified = {
        "Entity_ID": eid,
        "Entity_Type": etype.upper(),
        "Name": raw_name.title() if etype == "drug" else raw_name,
        "Grade_TradeName": abbrev if abbrev else "API (Pure)",
        "MW_g_mol": mw_val,
        "Tm_K": tm_val if tm_val is not None else "N/A",
        "Tg_10k_MC_Final_K": tg_val,
        "Density_10k_MC_Final_g_cm3": dens_val,
        "delta_D_10k_MC_Final_MPa0.5": dD_val,
        "delta_P_10k_MC_Final_MPa0.5": dP_val,
        "delta_H_10k_MC_Final_MPa0.5": dH_val,
        "delta_t_10k_MC_Final_MPa0.5": dt_val,
        "R0_MPa0.5": r0_val,
        "LogP": logp_val if logp_val is not None else "N/A",
        "TPSA_A2": tpsa_val if tpsa_val is not None else "N/A",
        "HBD": hbd_val if hbd_val is not None else "N/A",
        "HBA": hba_val if hba_val is not None else "N/A",
        "BCS_Class": bcs_val if bcs_val is not None else "N/A",
        "SMILES": str(smiles_val) if smiles_val else ""
    }
    all_rows.append(unified)
    
    if etype == "drug":
        drug_rows.append({
            "Entity_ID": eid,
            "Drug_Name": raw_name.title(),
            "Canonical_SMILES": str(smiles_val) if smiles_val else "",
            "MW_g_mol": mw_val,
            "Tm_Kelvin": tm_val,
            "Tg_10k_MC_Final_K": tg_val,
            "Density_10k_MC_Final_g_cm3": dens_val,
            "delta_D_10k_MC_Final_MPa0.5": dD_val,
            "delta_P_10k_MC_Final_MPa0.5": dP_val,
            "delta_H_10k_MC_Final_MPa0.5": dH_val,
            "delta_t_10k_MC_Final_MPa0.5": dt_val,
            "R0_MPa0.5": r0_val,
            "LogP": logp_val,
            "TPSA_A2": tpsa_val,
            "HBD": hbd_val,
            "HBA": hba_val,
            "BCS_Class": bcs_val
        })
    else:
        polymer_rows.append({
            "Entity_ID": eid,
            "Polymer_Name": raw_name,
            "Commercial_Grade": abbrev,
            "Repeat_Unit_SMILES": str(smiles_val) if smiles_val else "",
            "Tg_10k_MC_Final_K": tg_val,
            "Bulk_Density_10k_MC_Final_g_cm3": dens_val,
            "True_Density_Reference_g_cm3": true_dens,
            "delta_D_10k_MC_Final_MPa0.5": dD_val,
            "delta_P_10k_MC_Final_MPa0.5": dP_val,
            "delta_H_10k_MC_Final_MPa0.5": dH_val,
            "delta_t_10k_MC_Final_MPa0.5": dt_val,
            "R0_MPa0.5": r0_val,
            "Repeat_Unit_MW_g_mol": mw_val
        })

# 1. Write CSV
DOWNLOADS_DIR = r"C:\Users\Admin\Downloads"
DATA_STORE_DIR = r"C:\Users\Admin\.gemini\antigravity\scratch\pharmapolyscope-input-generator\data\store"

csv_filename = "PharmaPolySCOPE_10k_MC_Final_Inputs.csv"
csv_path_dl = os.path.join(DOWNLOADS_DIR, csv_filename)
csv_path_store = os.path.join(DATA_STORE_DIR, csv_filename)

headers_csv = list(all_rows[0].keys())
for p in [csv_path_dl, csv_path_store]:
    with open(p, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=headers_csv)
        writer.writeheader()
        writer.writerows(all_rows)
    print(f"Saved CSV: {p}")

# 2. Write Excel (.xlsx) with 3 styled sheets
wb = openpyxl.Workbook()

# Sheet 1: All Entities
ws1 = wb.active
ws1.title = "All_10k_MC_Final"

# Sheet 2: Drugs
ws2 = wb.create_sheet(title="Drugs_10k_MC_Final")

# Sheet 3: Polymers
ws3 = wb.create_sheet(title="Polymers_10k_MC_Final")

def style_sheet(ws, rows_dict_list, title_text, header_bg="1E3A8A"):
    # Header styling
    font_title = Font(name="Calibri", size=14, bold=True, color="1E3A8A")
    font_subtitle = Font(name="Calibri", size=10, italic=True, color="475569")
    font_hdr = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    fill_hdr = PatternFill(start_color=header_bg, end_color=header_bg, fill_type="solid")
    border_thin = Border(
        left=Side(style='thin', color="CBD5E1"),
        right=Side(style='thin', color="CBD5E1"),
        top=Side(style='thin', color="CBD5E1"),
        bottom=Side(style='thin', color="CBD5E1")
    )
    fill_alt = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    # Title Block
    ws.merge_cells("A1:K1")
    ws["A1"] = title_text
    ws["A1"].font = font_title
    
    ws.merge_cells("A2:K2")
    ws["A2"] = "PharmaPolySCOPE ASD Screening Specification • Exact 10k Monte Carlo Converged Final Inputs"
    ws["A2"].font = font_subtitle
    
    ws.append([]) # Row 3 blank
    
    headers = list(rows_dict_list[0].keys())
    ws.append(headers)
    hdr_row_idx = 4
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=hdr_row_idx, column=col_idx)
        cell.font = font_hdr
        cell.fill = fill_hdr
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_thin
    
    ws.row_dimensions[hdr_row_idx].height = 28
    
    for r_idx, row_data in enumerate(rows_dict_list, start=hdr_row_idx + 1):
        row_vals = [row_data[h] for h in headers]
        ws.append(row_vals)
        ws.row_dimensions[r_idx].height = 20
        is_alt = (r_idx % 2 == 0)
        for c_idx, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.border = border_thin
            if is_alt:
                cell.fill = fill_alt
            
            # Numeric formatting
            if isinstance(val, float):
                cell.number_format = "0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif isinstance(val, int):
                cell.number_format = "0"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif c_idx == 1: # Entity ID
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(name="Calibri", bold=True, color="0284C7")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row >= hdr_row_idx and cell.value:
                val_str = str(cell.value)
                if len(val_str) > max_len:
                    max_len = len(val_str)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)

style_sheet(ws1, all_rows, "PharmaPolySCOPE: All Formulation Entities (10k MC Final Inputs)", header_bg="1E3A8A")
style_sheet(ws2, drug_rows, "PharmaPolySCOPE: BCS Class II Drugs (10k MC Final Inputs)", header_bg="0F766E")
style_sheet(ws3, polymer_rows, "PharmaPolySCOPE: Commercial ASD Polymers (10k MC Final Inputs)", header_bg="4338CA")

excel_filename = "PharmaPolySCOPE_10k_MC_Final_Inputs.xlsx"
excel_path_dl = os.path.join(DOWNLOADS_DIR, excel_filename)
excel_path_store = os.path.join(DATA_STORE_DIR, excel_filename)

wb.save(excel_path_dl)
wb.save(excel_path_store)
print(f"Saved Excel: {excel_path_dl}")
print(f"Saved Excel: {excel_path_store}")

# 3. Also write JSON
json_filename = "PharmaPolySCOPE_10k_MC_Final_Inputs.json"
json_path_dl = os.path.join(DOWNLOADS_DIR, json_filename)
json_path_store = os.path.join(DATA_STORE_DIR, json_filename)

export_bundle = {
    "title": "PharmaPolySCOPE 10k Monte Carlo Final Converged Inputs",
    "description": "Ready-to-paste formulation parameters for small-molecule drugs and commercial polymer carriers after 10,000 Monte Carlo uncertainty iterations.",
    "software_target": "PharmaPolySCOPE ASD Formulation Screening Engine",
    "total_records": len(all_rows),
    "total_drugs": len(drug_rows),
    "total_polymers": len(polymer_rows),
    "data": all_rows
}

for jp in [json_path_dl, json_path_store]:
    with open(jp, "w", encoding="utf-8") as f:
        json.dump(export_bundle, f, indent=2)
    print(f"Saved JSON: {jp}")

print("\nALL EXPORT FILES GENERATED SUCCESSFULLY!")
